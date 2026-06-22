"""Deterministic XLSX generator (五期 B+ 多类型扩展).

The model supplies *structured worksheets* (each a headers + rows table); this
module renders them into a real ``.xlsx`` using a fixed, code-controlled style
(``openpyxl``). The model never writes rendering code (A 铁律).

Design goals mirror ``pptx_generator``: deterministic, dependency-light
(``openpyxl`` + stdlib), defensive (normalize loose / oversized output before
rendering), and generic / open-source safe (single env-overridable accent).

Public API:
    normalize_workbook(raw) -> dict     # validate + clamp into a canonical shape
    build_xlsx(workbook) -> bytes       # render canonical workbook → .xlsx bytes
    safe_filename(title, ext="xlsx")    # re-exported from file_gen_common
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import file_gen_common as common

# --- Limits (clamp model output) ----------------------------------------------
MAX_SHEETS = 12
MAX_COLS = 100
MAX_ROWS = 5000
MAX_CELL_CHARS = 2000
MAX_TITLE_CHARS = 120
_SHEET_NAME_MAX = 31  # Excel hard limit
_INVALID_SHEET_CHARS = set(r"[]:*?/\\")


def _safe_sheet_name(name: str, index: int, used: set[str]) -> str:
    """Excel sheet names: ≤31 chars, no ``[]:*?/\\``, unique, non-empty."""
    cleaned = "".join("_" if ch in _INVALID_SHEET_CHARS else ch for ch in (name or ""))
    cleaned = cleaned.strip().strip("'")[:_SHEET_NAME_MAX]
    if not cleaned:
        cleaned = f"Sheet{index}"
    base = cleaned
    n = 1
    while cleaned.lower() in used:
        suffix = f"_{n}"
        cleaned = base[: _SHEET_NAME_MAX - len(suffix)] + suffix
        n += 1
    used.add(cleaned.lower())
    return cleaned


def _normalize_sheet(raw: Any) -> dict[str, Any]:
    """One sheet → {name, headers: [str], rows: [[cell]]} (rectangular)."""
    if not isinstance(raw, dict):
        raw = {}
    name = common.clean_text(raw.get("name") or raw.get("title"), _SHEET_NAME_MAX)

    raw_headers = raw.get("headers") or raw.get("columns") or []
    if not isinstance(raw_headers, list):
        raw_headers = []
    headers = [common.clean_text(h, 200) for h in raw_headers][:MAX_COLS]

    raw_rows = raw.get("rows") or raw.get("data") or []
    if not isinstance(raw_rows, list):
        raw_rows = []

    # derive headers from dict-row keys when omitted
    if not headers and raw_rows and isinstance(raw_rows[0], dict):
        seen: list[str] = []
        for r in raw_rows[:MAX_ROWS]:
            if isinstance(r, dict):
                for k in r.keys():
                    kk = common.clean_text(k, 200)
                    if kk and kk not in seen:
                        seen.append(kk)
                        if len(seen) >= MAX_COLS:
                            break
        headers = seen

    width = len(headers)
    rows: list[list[Any]] = []
    for r in raw_rows:
        if isinstance(r, dict):
            cells = [common.as_cell(r.get(h, ""), MAX_CELL_CHARS) for h in headers]
        elif isinstance(r, list):
            src = r[:width] if width else r[:MAX_COLS]
            cells = [common.as_cell(c, MAX_CELL_CHARS) for c in src]
        else:
            cells = [common.as_cell(r, MAX_CELL_CHARS)]
        if width:
            cells = (cells + [""] * (width - len(cells)))[:width] if len(cells) < width else cells[:width]
        rows.append(cells)
        if len(rows) >= MAX_ROWS:
            break

    return {"name": name, "headers": headers, "rows": rows}


def normalize_workbook(raw: Any) -> dict[str, Any]:
    """Validate + clamp loose model output into a canonical workbook.

    Canonical shape::

        {"title": str, "sheets": [{"name": str, "headers": [str], "rows": [[cell]]}, ...]}

    Accepts a single-table payload (``headers``/``rows`` at top level) as a
    one-sheet workbook. Never raises — returns something renderable.
    """
    if not isinstance(raw, dict):
        raw = {}
    title = common.clean_text(raw.get("title") or raw.get("name"), MAX_TITLE_CHARS) or "工作簿"

    raw_sheets = raw.get("sheets")
    if not isinstance(raw_sheets, list) or not raw_sheets:
        # single-table payload at top level → one sheet
        if raw.get("headers") or raw.get("columns") or raw.get("rows") or raw.get("data"):
            raw_sheets = [raw]
        else:
            raw_sheets = []

    used: set[str] = set()
    sheets: list[dict[str, Any]] = []
    for i, s in enumerate(raw_sheets, start=1):
        sheet = _normalize_sheet(s)
        if not sheet["headers"] and not sheet["rows"]:
            continue
        sheet["name"] = _safe_sheet_name(sheet["name"], i, used)
        sheets.append(sheet)
        if len(sheets) >= MAX_SHEETS:
            break

    if not sheets:
        sheets = [{"name": _safe_sheet_name("", 1, used), "headers": ["列1"], "rows": [[""]]}]

    return {"title": title, "sheets": sheets}


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------


def _col_width(values: list[Any]) -> float:
    """Rough auto-width: widest cell (CJK counts ~2), clamped to a sane range."""
    longest = 0
    for v in values:
        text = "" if v is None else str(v)
        # CJK / wide chars take ~2 display columns
        w = sum(2 if ord(ch) > 0x2E7F else 1 for ch in text)
        longest = max(longest, w)
    return max(8.0, min(longest + 2.0, 60.0))


def build_xlsx(workbook: Any) -> bytes:
    """Render a (canonical or loose) workbook into ``.xlsx`` bytes."""
    workbook = normalize_workbook(workbook)
    accent = common.accent_hex()

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet; we add our own

    header_fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for sheet in workbook["sheets"]:
        ws = wb.create_sheet(title=sheet["name"])
        headers = sheet["headers"]
        rows = sheet["rows"]

        start_row = 1
        if headers:
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            ws.freeze_panes = "A2"  # keep header visible on scroll
            start_row = 2

        for r, row in enumerate(rows, start=start_row):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

        # column widths from header + sampled rows (cap sampling for big tables)
        ncols = len(headers) if headers else (len(rows[0]) if rows else 1)
        sample = rows[:200]
        for c in range(1, ncols + 1):
            col_values: list[Any] = []
            if headers and c <= len(headers):
                col_values.append(headers[c - 1])
            for row in sample:
                if c <= len(row):
                    col_values.append(row[c - 1])
            ws.column_dimensions[get_column_letter(c)].width = _col_width(col_values)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def safe_filename(title: str, ext: str = "xlsx") -> str:
    return common.safe_filename(title, ext, fallback="workbook")


if __name__ == "__main__":  # pragma: no cover — local smoke test
    import sys

    from openpyxl import load_workbook

    sample = {
        "title": "2024 区域经营",
        "sheets": [
            {
                "name": "区域汇总",
                "headers": ["区域", "销售额(万)", "门店数", "单店均值"],
                "rows": [
                    ["华东", 1240.5, 47, 26.4],
                    ["华南", 980, 33, 29.7],
                    {"区域": "华北", "销售额(万)": 610, "门店数": 21, "单店均值": 29.0},
                ],
            },
            {"name": "明细", "headers": ["门店", "城市", "营收"], "rows": [["A001", "上海", 88.2]]},
        ],
    }
    data = build_xlsx(sample)
    out = sys.argv[1] if len(sys.argv) > 1 else "/tmp/sample.xlsx"
    with open(out, "wb") as f:
        f.write(data)
    reopened = load_workbook(io.BytesIO(data))
    print(f"OK: {len(data)} bytes, sheets={reopened.sheetnames} → {out}")
    print(f"filename: {safe_filename(sample['title'])}")
